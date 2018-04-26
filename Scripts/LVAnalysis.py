from openpyxl import Workbook

def main ():

  wb = Workbook()

  # grab the active worksheet
  ws = wb.active

  # Data can be assigned directly to cells
  ws['A1'] = 42
  tempCell = 'A'
  tempCellNum = 2
  

  # Save the file
  



  

  compareLAA = open("/Bioinformatics_Project/Data/DatabaseNewLRegionAA.txt", "r") #Parsed sequence of L regions
  viterbi = open("/Bioinformatics_Project/Scripts/Output/LVViterbiOutput.txt", "r") #ViterbiOutput file to be analyzed
  output = open("/Bioinformatics_Project/Scripts/Output/LVSeqOutAnalysis.txt", "w")

  index = 1
  index2 = 1
  total_sequences = 0
  compareLDict = {}
  frequencies = {}
  total = 0
  last = ""
  tempID = ""
  temp = ""
  for line in compareLAA:
    """print(line)
    line = line.replace("\n","")
    input()"""
    if line[0] == ">":
      temp = ""
      for char in line:
        if char == ".":
          break
        elif char != ">":
          temp = temp + char
          
    else:
      line = line.rstrip("\n\t")
      compareLDict[temp] = line
      temp = ""

  """for key in compareLDict:
    print(key+ " " + str(compareLDict[key]))
    input()"""




      
  for line in viterbi:
    if index == 1:
      for char in line:
        if char == ".":
          break
        elif char != ">":
          tempID = tempID + char
      output.write(line)  
      index = 2
      total_sequences += 1
    elif index == 2:
      output.write(line)
      if tempID in compareLDict:
        seqTemp = ""
        tempLen = len(compareLDict[tempID])
        for x in range(tempLen):
          seqTemp = seqTemp + "L"
      for y in range(len(line) - tempLen - 1):
        seqTemp = seqTemp + "V"
      output.write(seqTemp + "\n")
      #print(seqTemp)
      #input()
      seqTemp = ""
      index = 3
    elif index == 3:
      output.write(line)
      tempNum = (tempLen - line.find("V"))/len(line)
      fancyNum = str(abs(round(tempNum,5)))
      output.write(fancyNum + "\n")
      fullCell = tempCell + str(tempCellNum)
      #print(fullCell)
      #input()
      ws[fullCell] = str(fancyNum)
      tempCellNum += 1
    
      #print(line)
      index = 1
      tempID = ""
      
  print(total_sequences)
  wb.save("viterbiPlot.xlsx")
main()






